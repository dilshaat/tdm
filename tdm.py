from openpyxl import Workbook
from openpyxl import load_workbook
import csv
import argparse
import os, sys


tdm_template = 'tdm.xlsx' # Empty Excel Template for TDMs
wb = load_workbook(tdm_template) # Loding the Excel Template to memory
files = os.listdir('.') # Listing all files in this current directory

# TDM data type converter: Converts obscure SAP name to understandable english data type names
# No Python or SAP database specific technical data type conversion is performed here
def convert_data_type_names(old_data_type):
    if old_data_type == 'LANG':
        return 'LANGUAGE'
    elif old_data_type == 'DATS':
        return 'DATE'
    elif old_data_type == 'NUMC':
        return 'NUMERIC TEXT'
    elif old_data_type == 'TIMS':
        return 'Time HHMMSS'
    elif old_data_type == 'LCHR':
        return 'LONG CHAR'
    elif old_data_type == 'RAW':
        return 'Byte Sequence'
    elif old_data_type == 'CLNT':
        return 'Client'
    elif old_data_type == 'DEC':
        return 'Calculation/Amount Field'
    else:
        return old_data_type


# Validating passed command line arguments
parser = argparse.ArgumentParser()
parser.add_argument('objectname', help='The SAP object name', choices=['proj', 'wbs', 'cost', 'compunit'])
parser.add_argument('csv_file', help='SAP object spacific csv file to parse in the same folder')
args = parser.parse_args()

# Verifying if the passed csv file exists in the current directory
if args.csv_file not in files:
	sys.exit('csv file is not in this directory, please place the relevant csv file in the same directory')

# Creating Sheet Names list for this particular SAP Object
sheet_names = [
	args.objectname.upper() + ' Master',
	args.objectname.upper() + ' Classification',
	args.objectname.upper() + ' Partner Function',
	args.objectname.upper() + ' Document',
	args.objectname.upper() + ' Long Text'
]

print(sheet_names)

# Assigning the created Sheet Names to the actual spreadsheet[tdm.xlsx]
if len(sheet_names) == len(wb.sheetnames):
	counter = 0
	for sheet in wb:
		sheet.title = sheet_names[counter]
		counter += 1
print(wb.sheetnames)

# Creating WorkSheet objects for: Master, Classification, Partner Function, Document and Long Text
ws_master = wb[sheet_names[0]]
ws_class = wb[sheet_names[1]]
ws_pf = wb[sheet_names[2]]
ws_doc = wb[sheet_names[3]]
wb_lt = wb[sheet_names[4]]
# Setting Master Worksheet as focused worksheet
wb.active = 0

# Excel Sheet row counter
ws_row_counter = 3

# Reading CSV file and Writing to final TDM file
with open(args.csv_file, 'r', newline='') as csv_file:
	csv_lines = csv.reader(csv_file, delimiter=",")
	for line in csv_lines:
		#print(line[0], line[1], line[2], line[3], line[4], line[5], line[6].lstrip('0'))
		ws_master.cell(row=ws_row_counter, column=2, value=line[0])
		ws_master.cell(row=ws_row_counter, column=3, value=line[1])
		ws_master.cell(row=ws_row_counter, column=4, value=line[2])
		ws_master.cell(row=ws_row_counter, column=5, value=line[3])
		ws_master.cell(row=ws_row_counter, column=6, value=line[4])
		ws_master.cell(row=ws_row_counter, column=7, value=convert_data_type_names(line[5]))
		ws_master.cell(row=ws_row_counter, column=8, value=line[6].lstrip('0'))
		ws_row_counter += 1

wb.save('tdm_test.xlsx')